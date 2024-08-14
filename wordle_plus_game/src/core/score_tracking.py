from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from wordle_plus_game.src.components.text import Text

class ScoreTracking:
    """
    A class for tracking and managing game scores using an Excel workbook.

    Attributes:
        workbook_path (str): Path to the Excel workbook file.
        workbook (openpyxl.Workbook): Loaded workbook object.
    """

    def __init__(self, workbook_path='wordle_plus_game/local_files/Score_Boards.xlsx'):
        self.workbook_path = workbook_path
        self.workbook = load_workbook(workbook_path)

    def save_score(self, game_mode, data, screen=None, font=None, x=None, y=None, bg_color=None, top_score_message="New Top Score!"):
        """
        Save a score to the specified sheet in the workbook and check for top score.

        Args:
            game_mode (str): The name of the sheet to save the score.
            data (dict): A dictionary containing the score data to save.
            screen (pygame.Surface, optional): The Pygame screen surface for displaying the top score message.
            font (pygame.font.Font, optional): The font to use for rendering the text.
            x (int, optional): X-coordinate for the text position.
            y (int, optional): Y-coordinate for the text position.
            bg_color (tuple, optional): The background color of the text message.
            top_score_message (str, optional): The message to display for a top score.
        """
        difficulty = data.get("Difficulty")
        score = data.get("Score")
        time = data.get("Time")

        # Only save the score if it's the first time the game is over or if it's a top score.
        
        if self._is_top_score(game_mode, difficulty, score):
            self._update_top_score(game_mode, difficulty, score, time)
            if screen and font and x is not None and y is not None:
                self._display_top_score_message(screen, font, x, y, bg_color, top_score_message)

        # Save the score only once
        sheet = self.workbook[game_mode]
        next_row = sheet.max_row + 1
        for i, (key, value) in enumerate(data.items(), start=1):
            sheet[f"{get_column_letter(i)}{next_row}"] = value
        self.workbook.save(self.workbook_path)

    def _is_top_score(self, game_mode, difficulty, score):
        """
        Check if a score is a top score for the given game_mode and difficulty.

        Args:
            game_mode (str): The game mode (e.g., "Classic Scores" or "Hangman Scores").
            difficulty (str): The difficulty level.
            score (int): The score to check.

        Returns:
            bool: True if the score is a top score, False otherwise.
        """
        sheet = self.workbook["Top Scores"]
        for row in range(2, sheet.max_row + 1):
            if sheet[f"A{row}"].value == game_mode and sheet[f"B{row}"].value == difficulty:
                current_top_score = sheet[f"C{row}"].value
                return score < current_top_score
        return True

    def _update_top_score(self, game_mode, difficulty, score, time):
        """
        Update the top score in the "Top Scores" sheet.

        Args:
            game_mode (str): The game mode (e.g., "Classic Scores" or "Hangman Scores").
            difficulty (str): The difficulty level.
            score (int): The new top score.
            time (str): The time when the score was achieved.
        """
        sheet = self.workbook["Top Scores"]
        for row in range(2, sheet.max_row + 1):
            if sheet[f"A{row}"].value == game_mode and sheet[f"B{row}"].value == difficulty:
                sheet[f"C{row}"] = score
                sheet[f"D{row}"] = time
                self.workbook.save(self.workbook_path)
                return
        # If no existing top score found, add a new row
        next_row = sheet.max_row + 1
        sheet[f"A{next_row}"] = game_mode
        sheet[f"B{next_row}"] = difficulty
        sheet[f"C{next_row}"] = score
        sheet[f"D{next_row}"] = time
        self.workbook.save(self.workbook_path)

    def _display_top_score_message(self, screen, font, x, y, bg_color, message):
        """
        Display a top score message on the Pygame screen.

        Args:
            screen (pygame.Surface): The Pygame screen surface.
            font (pygame.font.Font): The font to use for rendering the text.
            x (int): X-coordinate for the text position.
            y (int): Y-coordinate for the text position.
            bg_color (tuple): The background color of the text message.
            message (str): The message to display.
        """
        top_score_text = Text(message, font, x, y, bg_color=bg_color)
        top_score_text.draw(screen)

    def load_scores(self, sheet_name):
        """
        Load scores from a specific sheet in the workbook.

        Args:
            sheet_name (str): The name of the sheet to load the scores from.

        Returns:
            list of dict: A list of dictionaries where each dictionary represents a row of score data.
        """
        sheet = self.workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[0]
        return [dict(zip(headers, row)) for row in rows[1:]]
